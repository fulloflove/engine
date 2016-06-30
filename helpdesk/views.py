# coding: utf-8
from django.shortcuts import render, get_object_or_404
from django.core.urlresolvers import reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpResponseRedirect, Http404
from forms import UserForm, IssueForm, ControlForm, CommentForm, DescriptionForm, ReportForm, IssuePeriodForm, \
    FilterForm
from django.conf import settings
from models import Issue, Project, ServiceType, Status
from regions.models import Region
from regions.forms import ContactForm
from datetime import date, timedelta
from django.utils.translation import ugettext as _
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from tasks import issue_creation, issue_assign


def pagination(request, list_for_pagination):
    paginator = Paginator(list_for_pagination, settings.ENTRIES_PER_PAGE)
    page = request.GET.get('page')
    try:
        list_for_pagination = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        list_for_pagination = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        list_for_pagination = paginator.page(paginator.num_pages)
    return list_for_pagination


def get_queries(request):
    queries_without_page = request.GET.copy()
    if 'page' in queries_without_page:
        del queries_without_page['page']
    if 'view' in queries_without_page:
        del queries_without_page['view']
    return queries_without_page


def get_pagination_range(list_for_pagination, pages_limit=10, adj_pages=3):
    first_break = False
    last_break = False
    if list_for_pagination.paginator.num_pages < pages_limit:
        pagination_range = range(2, list_for_pagination.paginator.num_pages)
    elif adj_pages + 2 < list_for_pagination.number < (list_for_pagination.paginator.num_pages - adj_pages - 1):
        pagination_range = range(list_for_pagination.number-adj_pages, list_for_pagination.number+adj_pages+1)
        first_break = True
        last_break = True
    elif list_for_pagination.number > adj_pages + 2:
        pagination_range = range(list_for_pagination.number-adj_pages, list_for_pagination.paginator.num_pages)
        first_break = True
    else:
        pagination_range = range(2, list_for_pagination.number+adj_pages+1)
        last_break = True
    return {'pagination_range': pagination_range, 'first_break': first_break, 'last_break': last_break}


@login_required
def issues(request, context):
    context['issues'] = pagination(request, context['issues'])
    context.update(get_pagination_range(list_for_pagination=context['issues']))
    return render(request, 'helpdesk/issues.html', context)


def issues_dates_filter(issue_list, **kwargs):
    filter_dict = {}
    for (k, v) in kwargs.iteritems():
        if k.endswith('start') and v != '':
            key = k.replace('_start', '')
            if key in Issue._meta.get_all_field_names():
                filter_dict[key + '__gte'] = v
        elif k.endswith('end') and v != '':
            key = k.replace('_end', '')
            if key in Issue._meta.get_all_field_names():
                filter_dict[key + '__lte'] = v
    issue_list = issue_list.filter(**filter_dict)
    return issue_list


def issues_period(request, context):
    period_form = IssuePeriodForm()
    issue_list = []

    if 'period_form' in request.GET:
        period_form = IssuePeriodForm(request.GET)
        if period_form.is_valid():
            issue_list = context['issues']
            #issue_list = issues_dates_filter(issue_list, period_form.cleaned_data)
            if period_form.cleaned_data.get('start'):
                start_date = period_form.cleaned_data.get('start')
                issue_list = issue_list.filter(opened__gte=start_date)
            if period_form.cleaned_data.get('end'):
                end_date = period_form.cleaned_data.get('end')
                issue_list = issue_list.filter(opened__lte=end_date)
            context['queries'] = get_queries(request)

    context['period_form'] = period_form
    context['issues'] = issue_list
    context['title'] = _('Found issues')
    return issues(request, context)


def open_issues(context):
    context.update({'issues': Issue.objects.exclude(status=4), 'title': _('Open issues')})
    return context


def closed_issues(context):
    context.update({'issues': Issue.objects.filter(status=4), 'title': _('Closed issues')})
    return context


def any_issues(context):
    context.update({'issues': Issue.objects.all(), 'title': _('Any issues')})
    return context

@login_required
def my_issues(request, context):
    context['issues'] = context['issues'].filter(assignee=request.user)
    return context


def issues_main_view(request, active_nav, item_id=None, status=None, scope=None):
    context = {'active_nav': active_nav, 'status': status}
    if status is None:
        context['status'] = 'open'
        status_func = open_issues(context)
    elif status == 'any':
        status_func = any_issues(context)
    elif status == 'closed':
        status_func = closed_issues(context)
    else:
        raise Http404

    if scope == 'period':
        return_func = issues_period
    else:
        return_func = issues

    if active_nav == 'my':
        context = my_issues(request, status_func)
    elif active_nav == 'all':
        context = status_func
    elif active_nav == 'project':
        context = issues_by_project(status_func, item_id)
    elif active_nav == 'service_type':
        context = issues_by_service_type(status_func, item_id)
    elif active_nav == 'region':
        context = issues_by_region(status_func, item_id)
    elif active_nav == 'user':
        context = issues_by_assignee(status_func, item_id)
    else:
        raise Http404

    return return_func(request, context)


def issues_by_region(context, region_id):
    context['active_item'] = get_object_or_404(Region, pk=region_id)
    context['issues'] = context['issues'].filter(region=context['active_item'])
    return context


def issues_by_project(context, project_id):
    context['active_item'] = get_object_or_404(Project, pk=project_id)
    context['issues'] = context['issues'].filter(project=context['active_item'])
    return context


def issues_by_service_type(context, service_type_id):
    context['active_item'] = get_object_or_404(ServiceType, pk=service_type_id)
    context['issues'] = context['issues'].filter(service_type=context['active_item'])
    return context


def issues_by_assignee(context, assignee_id):
    context['active_item'] = get_object_or_404(User, pk=assignee_id)
    context['issues'] = context['issues'].filter(assignee=context['active_item'])
    return context


@login_required
def issue(request, issue_id, context=None):
    if context is None:
        context = {}
    current_issue = get_object_or_404(Issue, pk=issue_id)
    comment_form = CommentForm()
    description_form = DescriptionForm(instance=current_issue)
    report_form = ReportForm(instance=current_issue)
    control_form = ControlForm(instance=current_issue)
    if request.method == 'POST':
        if 'comment_form' in request.POST:
            comment_form = CommentForm(request.POST)
            if comment_form.is_valid():
                curr_comment = comment_form.save(commit=False)
                curr_comment.author = request.user
                curr_comment.issue = current_issue
                curr_comment.save()
                comment_form = CommentForm()
        if 'description_form' in request.POST:
            description_form = DescriptionForm(request.POST, instance=current_issue)
            if description_form.is_valid():
                description_form.save()
                description_form = DescriptionForm(instance=current_issue)
        if 'report_form' in request.POST:
            report_form = ReportForm(request.POST, instance=current_issue)
            if report_form.is_valid():
                report_form.save()
                report_form = ReportForm(instance=current_issue)
        if 'control_form' in request.POST:
            control_form = ControlForm(request.POST, instance=current_issue)
            if control_form.is_valid():
                control_form.save()
                control_form = ControlForm(instance=current_issue)

    context['issue'] = current_issue
    context['comment_form'] = comment_form
    context['description_form'] = description_form
    context['report_form'] = report_form
    context['control_form'] = control_form
    context['status_list'] = Status.objects.all()
    return render(request, 'helpdesk/issue.html', context)


def issue_status(request, issue_id, status_id, context=None):
    if context is None:
        context = {}
    new_status = get_object_or_404(Status, pk=status_id)
    current_issue = get_object_or_404(Issue, pk=issue_id)
    if current_issue.status == new_status:
        message = "%s %s" % (_('Issue status is already set to: '), new_status)
    elif new_status == Status.objects.get(pk=4) and not current_issue.formed:
        message = _('Issue must be formed to be closed')
    elif new_status == Status.objects.get(pk=4) and not current_issue.contracts:
        message = _('Contracts must be set for issue to be closed')
    else:
        current_issue.status = new_status
        current_issue.save()
        message = "%s %s" % (_('Issue status is changed. New status: '), new_status)
    context['status_message'] = message
    return issue(request, issue_id, context)


def save_issue(form, current_user, continue_editing=False):
    curr_issue = form.save(commit=False)
    curr_region = curr_issue.region
    curr_region.region_count += 1
    curr_region.save()
    curr_issue.issue_id = 'FU-{0}-{1}-{2}'.format(curr_region.region_id,
                                                  date.today().strftime("%y"),
                                                  curr_region.region_count)
    curr_issue.creator = current_user
    curr_issue.save()
    form.save_m2m()
    issue_creation.delay(curr_issue)

    if continue_editing:
        return HttpResponseRedirect(reverse('helpdesk:edit_issue', args=[curr_issue.id]))
    return HttpResponseRedirect(reverse('helpdesk:issue', args=[curr_issue.id]))


def row_for_export(issue_obj):
    from django.utils.encoding import smart_str

    fields = [
        issue_obj.issue_id,
        issue_obj.external_id,
        issue_obj.service_type.name_short,
        issue_obj.report_description,
        issue_obj.report_solution,
        issue_obj.opened,
        issue_obj.control,
        issue_obj.region.gu,
    ]

    fields_smart_str = [smart_str(v) for v in fields]

    return fields_smart_str


def export_issues(request, export_type, issue_list=None):
    from django.http import HttpResponse
    #queries_clean = {k: v for k, v in queries.iteritems() if v and k != 'export'}
    if issue_list:
        if export_type == 'csv':
            import csv
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="export.csv"'
            writer = csv.writer(response, delimiter=';')
            response.write(u'\ufeff'.encode('utf8'))
            for obj in issue_list:
                writer.writerow(row_for_export(obj))
            return response
        elif export_type == 'xls':
            import openpyxl
            from openpyxl.cell import get_column_letter
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="export.xlsx"'

            wb = openpyxl.Workbook()
            ws = wb.get_active_sheet()
            ws.title = "Issues"

            row_num = 0
            columns = (15, 20, 10, 70, 70, 15, 15, 20)
            for col_num in xrange(len(columns)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num]

            for obj in issue_list:
                row_num += 1
                row = row_for_export(obj)
                for col_num in xrange(len(row)):
                    c = ws.cell(row=row_num, column=col_num + 1)
                    c.value = row[col_num]
                    c.style.alignment.wrap_text = True
            wb.save(response)
            return response
        elif export_type == 'html':
            return render(request, 'helpdesk/filter_printable.html', {'issues': issue_list.order_by('opened')})
        else:
            raise Http404
    raise Http404


@login_required
def filter_issues(request):
    filter_form = FilterForm(request.GET)
    context = {'filter_form': filter_form}
    key_filters = ('project',
                   'service_type',
                   'status',
                   'priority',
                   'region',
                   'source',
                   'assignee',
                   'creator',
                   'component',
                   )

    date_filters = ('opened_start', 'opened_end', 'control_start', 'control_end')

    m2m_filters = ('contracts',)

    context['filtered'] = False
    if 'search' in request.GET:
        if filter_form.is_valid():
            context['queries'] = get_queries(request)
            context['filtered'] = True
            filtered_issues = Issue.objects.all()
            date_filters_dict = {}
            params = [v for v in request.GET.keys() if request.GET.get(v) != '' and v != 'page']
            for p in params:
                if p in key_filters:
                    qs_str = p + '__in'
                    filtered_issues = filtered_issues.filter(**{qs_str: request.GET.getlist(p)})
                elif p in m2m_filters:
                    qs_str = p + '__in'
                    filtered_issues = filtered_issues.filter(**{qs_str: request.GET.getlist(p)}).distinct()
                elif p in date_filters:
                    date_filters_dict[p] = filter_form.cleaned_data.get(p)
            filtered_issues = issues_dates_filter(filtered_issues, **date_filters_dict)

            if 'export' in request.GET:
                return export_issues(request, request.GET.get('export'), filtered_issues)

            context['issues'] = pagination(request, filtered_issues)
            context.update(get_pagination_range(list_for_pagination=context['issues']))
            context['title'] = _('Found issues')

            if 'view' in request.GET:
                context['extended_view'] = True

    return render(request, 'helpdesk/filter.html', context)


@login_required
def new_issue(request):
    current_user = request.user
    context = {}
    contact_form = ContactForm()

    if request.method == 'POST' and ('new_issue' in request.POST or 'continue_editing' in request.POST):
        form = IssueForm(request.POST)
        if form.is_valid():
            if 'continue_editing' in request.POST:
                return save_issue(form, current_user, continue_editing=True)
            return save_issue(form, current_user)

    elif request.method == 'POST' and 'new_contact' in request.POST:
        contact_form = ContactForm(request.POST)     
        if contact_form.is_valid():
            curr_region = get_object_or_404(Region, pk=contact_form['region'].value())
            contact = contact_form.save(commit=False)
            contact.region = curr_region
            contact.save()
            contact_form = ContactForm()
            submitted_data = request.POST.copy()
            submitted_data['contact'] = contact.id
            form = IssueForm(submitted_data)
            context['created_contact'] = contact
        else:
            form = IssueForm(request.POST)

    else:
        initial_data = {'opened': date.today(),
                        'control': date.today() + timedelta(days=7),
                        'assignee': current_user.id,
                        'project': 1,
                        'service_type': 4,
                        'status': 1,
                        'priority': 3,
                        'formed': False}
        form = IssueForm(initial=initial_data)

    context['form'] = form
    context['contact_form'] = contact_form
    return render(request, 'helpdesk/new_issue.html', context)


@login_required
def edit_issue(request, issue_id):
    current_issue = get_object_or_404(Issue, pk=issue_id)
    contact_form = ContactForm()
    form = IssueForm(instance=current_issue)

    if request.method == 'POST' and ('new_issue' in request.POST or 'continue_editing' in request.POST):
        current_assignee = current_issue.assignee
        form = IssueForm(request.POST or None, instance=current_issue)
        if form.is_valid():
            form.save()
            if current_assignee != current_issue.assignee:
                issue_assign.delay(current_issue)
            if 'continue_editing' in request.POST:
                return HttpResponseRedirect(reverse('helpdesk:edit_issue', args=[current_issue.id]))
            return HttpResponseRedirect(reverse('helpdesk:issue', args=[current_issue.id]))

    elif request.method == 'POST' and 'new_contact' in request.POST:
        contact_form = ContactForm(request.POST)
        if contact_form.is_valid():
            new_contact = contact_form.save(commit=False)
            new_contact.region = current_issue.region
            new_contact.save()
            contact_form = ContactForm()
            current_issue.contact = new_contact
            current_issue.save()
            form = IssueForm(instance=current_issue)

    return render(request, 'helpdesk/new_issue.html', {'form': form,
                                                       'current_issue': current_issue,
                                                       'contact_form': contact_form})


@login_required
def popover(request, region_id=None):
    if region_id:
        curr_region = get_object_or_404(Region, pk=region_id)
    else:
        curr_region = None
    return render(request, 'helpdesk/popover.html', {'region': curr_region})


@login_required
def account_change(request):
    success = False
    if request.method == 'POST':
        user_form = UserForm(data=request.POST, instance=request.user)
        if user_form.is_valid():
            user_form.save()
            success = True
    else:
        user_form = UserForm(instance=request.user)

    return render(request, 'registration/edit_account.html', {'user_form': user_form, 'success': success})
